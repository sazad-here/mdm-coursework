"""Verify Assignment 4 workbooks contain required structure and formulas."""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
DATA = BASE / "data"
ok, fail = [], []


def check(cond: bool, msg: str):
    (ok if cond else fail).append(msg)


# --- P03_02 ---
wb = openpyxl.load_workbook(DATA / "P03_02.xlsx", data_only=False)
check("Data" in wb.sheetnames, "P03_02: sheet Data exists")
ws = wb["Data"]
check(str(ws["H1"].value) == "Salary_Group", "P03_02: Salary_Group in H1")
check("PERCENTILE" not in str(ws["H2"].value or ""), "P03_02: H2 not tertiles (uses $40K bands)")
check("40000" in str(ws["H2"].value or ""), "P03_02: H2 uses $40K breakpoint")

for sn in ("P2a_Gender", "P2b_Age", "P2c_Salary"):
    check(sn in wb.sheetnames, f"P03_02: sheet {sn} exists")
    s = wb[sn]
    found = False
    for row in s.iter_rows(max_row=25, max_col=6, values_only=True):
        for v in row:
            if v and "COUNTIFS" in str(v):
                found = True
                break
    check(found, f"P03_02: {sn} contains COUNTIFS")

check("P34_RowPct" in wb.sheetnames and "P34_ColPct" in wb.sheetnames, "P03_02: P34 sheets exist")

p2a = wb["P2a_Gender"]
has_chart = len(getattr(p2a, "_charts", [])) >= 1
check(has_chart, "P03_02: P2a_Gender has embedded chart")

wb_d = openpyxl.load_workbook(DATA / "P03_02.xlsx", data_only=True)
d = wb_d["Data"]
male_total = female_total = 0
male_counts = [0] * 5
female_counts = [0] * 5
op_to_i = {
    "Strongly disagree": 0,
    "Disagree": 1,
    "Neutral": 2,
    "Agree": 3,
    "Strongly agree": 4,
}
for r in range(2, d.max_row + 1):
    g = d.cell(r, 3).value
    op = d.cell(r, 7).value
    i = op_to_i.get(op)
    if i is None:
        continue
    if g == 1:
        male_total += 1
        male_counts[i] += 1
    elif g == 2:
        female_total += 1
        female_counts[i] += 1
male_pct_sum = sum(100.0 * c / male_total for c in male_counts) if male_total else 0.0
female_pct_sum = sum(100.0 * c / female_total for c in female_counts) if female_total else 0.0
check(
    abs(male_pct_sum - 100) < 0.01 and abs(female_pct_sum - 100) < 0.01,
    "P03_02: P2a % columns sum to 100",
)

# --- P03_08 ---
wb8 = openpyxl.load_workbook(DATA / "P03_08.xlsx", data_only=False)
check(wb8["Data"]["A1"].value == "Employee", "P03_08: Data A1 = Employee (not corrupted)")
check("P8_Summary" in wb8.sheetnames, "P03_08: P8_Summary exists")
check("P21_Corr" not in wb8.sheetnames, "P03_08: no stray P21_Corr")
check("P21_Charts" not in wb8.sheetnames, "P03_08: no stray P21_Charts")
d8 = wb8["Data"]
check("Edu_Recode" in str(d8["I1"].value), "P03_08: Edu_Recode in I1")
check("Middle-Aged" in str(d8["J2"].value or ""), "P03_08: J2 uses Middle-Aged (rubric spelling)")
check("AVERAGEIFS(" in str(wb8["P8_Summary"]["B4"].value or ""), "P03_08: P8_Summary mean uses AVERAGEIFS")
med_formula = str(wb8["P8_Summary"]["C4"].value or "").upper()
check("MEDIAN(" in med_formula and "FILTER(" in med_formula, "P03_08: P8_Summary median uses MEDIAN/FILTER")
check("STDEV.S(" in str(wb8["P8_Summary"]["D4"].value or ""), "P03_08: P8_Summary stdev uses STDEV.S/FILTER")

# --- P03_21 ---
wb21 = openpyxl.load_workbook(DATA / "P03_21.xlsx", data_only=False)
check("P21_Answers" in wb21.sheetnames, "P03_21: P21_Answers exists (interactive regression)")
check("P21_Corr" in wb21.sheetnames and "P21_Charts" in wb21.sheetnames, "P03_21: solution sheets exist")
ans = wb21["P21_Answers"]
check("LINEST" in str(ans["B5"].value or ""), "P03_21: P21_Answers R² uses LINEST")
c = wb21["P21_Corr"]
check("CORREL" in str(c["B10"].value or ""), "P03_21: P21_Corr has CORREL")
check("LINEST" in str(c["N10"].value or ""), "P03_21: LINEST slopes in column N")
check("LINEST" in str(c["O10"].value or ""), "P03_21: LINEST intercepts in column O")
check("ABS(B10)" in str(c["C10"].value or ""), "P03_21: |r| helper uses ABS")
check("MAX(C10:C13)" in str(c["B14"].value or ""), "P03_21: strongest predictor formula uses helper |r|")
ch = wb21["P21_Charts"]
n_charts = len(getattr(ch, "_charts", []))
check(n_charts >= 3, f"P03_21: P21_Charts has >=3 charts (found {n_charts})")

print("=== ASSIGNMENT 4 VERIFICATION ===\n")
print("PASSED (%d):" % len(ok))
for m in ok:
    print("  [OK]", m)
print("\nFAILED (%d):" % len(fail))
for m in fail:
    print("  [!!]", m)
print("\nResult:", "ALL CHECKS PASSED" if not fail else "SOME CHECKS FAILED")
raise SystemExit(1 if fail else 0)
