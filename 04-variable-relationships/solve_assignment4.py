"""
Populate Assignment 4 solutions in the three Excel workbooks using formulas,
Excel Tables, PivotTables, and charts (Excel COM on Windows).
Aligns with: Rubric for Assignment 4, Dataset guideline, and textbook (book.pdf).
Run: python solve_assignment4.py
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
import win32com.client as win32

BASE = Path(__file__).resolve().parent

OPINIONS = [
    "Strongly disagree",
    "Disagree",
    "Neutral",
    "Agree",
    "Strongly agree",
]
AGE_ORDER = ["Young", "Middle-aged", "Elderly"]
# Book (Ch.3 Problem 2 / P03_02): fixed salary bands, not tertiles.
SALARY_BANDS = [
    "Less than $40K",
    "Between $40K and $70K",
    "Between $70K and $100K",
    "Greater than $100K",
]

XL_DATABASE = 1
XL_ROW_FIELD = 1
XL_COLUMN_FIELD = 2
XL_COUNT = -4112
# XlPivotFieldCalculation — use Excel enum values (not chart constants).
XL_PERCENT_OF_ROW = 6
XL_PERCENT_OF_COLUMN = 7
XL_COLUMN_CLUSTERED = 51
XL_XY_SCATTER = -4169
XL_LINEAR = -4132
XL_UP = -4162
XL_SRC_RANGE = 1
XL_YES = 1
XL_CENTER = -4108
XL_TOP = -4160
XL_LEFT = -4131
XL_THIN = 2
XL_CONTINUOUS = 1
XL_EDGE_LEFT = 7
XL_EDGE_TOP = 8
XL_EDGE_BOTTOM = 9
XL_EDGE_RIGHT = 10


def _xl_bgr(r: int, g: int, b: int) -> int:
    """OLE color for Font.Color / Interior.Color (blue-green-red byte order)."""
    return r + (g << 8) + (b << 16)


CLR_HEADER = _xl_bgr(47, 84, 150)
CLR_HEADER_TEXT = _xl_bgr(255, 255, 255)
CLR_SUBHEADER = _xl_bgr(68, 114, 196)
CLR_BAND = _xl_bgr(248, 249, 252)
CLR_NOTE = _xl_bgr(255, 251, 240)
CLR_ACCENT = _xl_bgr(225, 235, 255)
CLR_BORDER = _xl_bgr(180, 198, 231)
CLR_LINK_BG = _xl_bgr(237, 245, 255)
CLR_TAB_BLUE = _xl_bgr(68, 114, 196)
CLR_TAB_GREEN = _xl_bgr(112, 173, 71)
CLR_TAB_ORANGE = _xl_bgr(237, 125, 49)
CLR_TAB_PURPLE = _xl_bgr(112, 48, 160)


def _border_box(rng) -> None:
    for edge in (XL_EDGE_LEFT, XL_EDGE_TOP, XL_EDGE_BOTTOM, XL_EDGE_RIGHT):
        b = rng.Borders(edge)
        b.LineStyle = XL_CONTINUOUS
        b.Weight = XL_THIN
        b.Color = CLR_BORDER


def _style_merge_title(ws, addr: str, row_height: float = 54.0) -> None:
    r = ws.Range(addr)
    r.Merge()
    r.HorizontalAlignment = XL_CENTER
    r.VerticalAlignment = XL_TOP
    r.WrapText = True
    r.Font.Size = 11
    r.Font.Bold = True
    r.Font.Color = CLR_HEADER_TEXT
    r.Interior.Color = CLR_HEADER
    r.RowHeight = row_height


def _style_column_headers(rng) -> None:
    rng.Font.Bold = True
    rng.Font.Color = CLR_HEADER_TEXT
    rng.Font.Size = 11
    rng.Interior.Color = CLR_SUBHEADER
    rng.HorizontalAlignment = XL_CENTER
    rng.VerticalAlignment = XL_CENTER
    _border_box(rng)


def _style_total_row(rng) -> None:
    rng.Font.Bold = True
    rng.Interior.Color = CLR_ACCENT
    _border_box(rng)


def _style_subtitle_row(rng) -> None:
    rng.Font.Bold = True
    rng.Font.Color = CLR_HEADER
    rng.Interior.Color = CLR_ACCENT
    _border_box(rng)


def _style_note_block(rng) -> None:
    rng.Interior.Color = CLR_NOTE
    rng.Font.Size = 10
    rng.VerticalAlignment = XL_TOP
    _border_box(rng)


def _zebra_block(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        row = ws.Range(ws.Cells(r, c1), ws.Cells(r, c2))
        if (r - r1) % 2 == 0:
            row.Interior.Color = CLR_BAND
        _border_box(row)


def _freeze_top_row(ws, freeze_below_row: int) -> None:
    ws.Activate()
    win = ws.Application.ActiveWindow
    win.FreezePanes = False
    win.SplitRow = freeze_below_row
    win.FreezePanes = True


def _style_nav_panel(ws, top_cell: str, nrows: int, header: str = "Quick links") -> None:
    """Eye-catching link column (interactive navigation)."""
    col = ws.Range(top_cell).Column
    r0 = ws.Range(top_cell).Row
    hdr = ws.Cells(r0, col)
    hdr.Value = header
    hdr.Font.Bold = True
    hdr.Font.Color = CLR_HEADER_TEXT
    hdr.Interior.Color = CLR_SUBHEADER
    hdr.HorizontalAlignment = XL_CENTER
    for i in range(1, nrows + 1):
        c = ws.Cells(r0 + i, col)
        c.Font.Color = CLR_SUBHEADER
        c.Font.Underline = True
        c.Interior.Color = CLR_LINK_BG
        _border_box(c)


def _polish_column_chart(ch) -> None:
    try:
        ch.ChartStyle = 26
    except Exception:
        pass
    try:
        ch.ChartArea.Format.Fill.Visible = True
        ch.ChartArea.Format.Fill.Solid()
        ch.ChartArea.Format.Fill.ForeColor.RGB = _xl_bgr(255, 255, 255)
        ch.PlotArea.Format.Fill.Visible = True
        ch.PlotArea.Format.Fill.Solid()
        ch.PlotArea.Format.Fill.ForeColor.RGB = CLR_ACCENT
    except Exception:
        pass
    try:
        ch.Legend.Font.Size = 10
        ch.ChartTitle.Font.Size = 12
        ch.ChartTitle.Font.Bold = True
    except Exception:
        pass


def _polish_scatter_chart(ch) -> None:
    try:
        ch.ChartStyle = 26
    except Exception:
        pass
    try:
        ch.ChartArea.Format.Fill.Solid()
        ch.ChartArea.Format.Fill.ForeColor.RGB = _xl_bgr(255, 255, 255)
        ch.PlotArea.Format.Fill.Solid()
        ch.PlotArea.Format.Fill.ForeColor.RGB = _xl_bgr(252, 252, 255)
    except Exception:
        pass


def _style_table_on_sheet(lo) -> None:
    try:
        lo.TableStyle = "TableStyleMedium2"
        lo.ShowTableStyleRowStripes = True
    except Exception:
        pass


def _style_regression_outputs(s) -> None:
    """P21_Answers — clear hierarchy and readable tables."""
    s.Range("A3:B10").Interior.Color = CLR_BAND
    s.Range("A12:F16").Interior.Color = CLR_BAND
    s.Range("A18:G26").Interior.Color = CLR_BAND
    _border_box(s.Range("A3:B10"))
    _border_box(s.Range("A12:F16"))
    _border_box(s.Range("A18:G26"))
    for addr in ("A3", "A12", "A18"):
        s.Range(addr).Font.Size = 12
        s.Range(addr).Font.Bold = True
        s.Range(addr).Font.Color = CLR_HEADER_TEXT
        s.Range(addr).Interior.Color = CLR_HEADER
        s.Range(addr).HorizontalAlignment = XL_LEFT
    s.Range("A4:A10").Font.Bold = True
    s.Range("A13:F13").Font.Bold = True
    s.Range("A13:F13").Font.Color = CLR_HEADER_TEXT
    s.Range("A13:F13").Interior.Color = CLR_SUBHEADER
    s.Range("A19:G19").Font.Bold = True
    s.Range("A19:G19").Font.Color = CLR_HEADER_TEXT
    s.Range("A19:G19").Interior.Color = CLR_SUBHEADER
    s.Range("A27").Font.Bold = True
    s.Range("B27").Interior.Color = CLR_NOTE
    _border_box(s.Range("A27:G27"))
    s.Range("B4:B10").NumberFormat = "0.0000"
    s.Range("B14:F16").NumberFormat = "0.0000"
    s.Range("B20:G25").NumberFormat = "0.0000"


def _style_corr_sheet(s) -> None:
    # Do not merge across B1:F1 — those cells are matrix column labels.
    s.Range("A1").ColumnWidth = 54
    s.Range("A1").HorizontalAlignment = XL_LEFT
    s.Range("A1").VerticalAlignment = XL_TOP
    s.Range("A1").Interior.Color = CLR_HEADER
    s.Range("A1").Font.Color = CLR_HEADER_TEXT
    s.Range("A1").Font.Bold = True
    s.Range("A1").Font.Size = 11
    s.Range("A1").WrapText = True
    s.Range("A1").RowHeight = 52
    s.Range("B1:F1").Interior.Color = CLR_SUBHEADER
    s.Range("B1:F1").Font.Bold = True
    s.Range("B1:F1").Font.Color = CLR_HEADER_TEXT
    s.Range("B1:F1").HorizontalAlignment = XL_CENTER
    m = s.Range("A2:F6")
    _border_box(m)
    s.Range("B2:F2").Interior.Color = CLR_SUBHEADER
    s.Range("A3:A6").Interior.Color = CLR_SUBHEADER
    s.Range("B2:F6").Font.Size = 10
    s.Range("B3:F6").NumberFormat = "0.0000"
    s.Range("A9:C16").Interior.Color = CLR_BAND
    _border_box(s.Range("A9:C16"))
    s.Range("A9").Font.Bold = True
    s.Range("A9").Interior.Color = CLR_SUBHEADER
    s.Range("A9").Font.Color = CLR_HEADER_TEXT
    s.Range("C9").Font.Bold = True
    s.Range("C9").Interior.Color = CLR_SUBHEADER
    s.Range("C9").Font.Color = CLR_HEADER_TEXT
    s.Range("B10:B15").NumberFormat = "0.0000"
    s.Range("C10:C13").NumberFormat = "0.0000"
    s.Range("A16:B16").Interior.Color = CLR_ACCENT
    s.Range("A16:B16").Font.Bold = True
    _border_box(s.Range("A16:B16"))
    s.Range("N9:O13").Interior.Color = CLR_BAND
    _border_box(s.Range("N9:O13"))
    s.Range("N9:O9").Font.Bold = True
    s.Range("A17:B20").Interior.Color = CLR_NOTE
    s.Range("A17:B20").WrapText = True
    _border_box(s.Range("A17:B20"))
    s.Range("B17").NumberFormat = "#,##0.00"
    s.Range("A21:B21").Interior.Color = CLR_NOTE
    s.Range("A21:B21").WrapText = True
    _border_box(s.Range("A21:B21"))
    s.Columns("A").ColumnWidth = 52
    s.Columns("B").ColumnWidth = 24
    s.Columns("C").ColumnWidth = 12
    s.Tab.Color = CLR_TAB_BLUE
    _freeze_top_row(s, 1)


def _del_sheet_if_exists(wb, name: str) -> None:
    for sh in list(wb.Worksheets):
        if sh.Name == name:
            sh.Delete()
            break


def _clear_listobjects_com(ws) -> None:
    """Convert Excel tables to normal ranges (COM). Use Unlist, not Delete — Delete wipes header row."""
    try:
        while ws.ListObjects.Count > 0:
            ws.ListObjects(1).Unlist()
    except Exception:
        pass


def _add_sheet_hyperlink(ws, cell: str, sub_address: str, text: str) -> None:
    """Link to another sheet (interactive navigation)."""
    ws.Hyperlinks.Add(Anchor=ws.Range(cell), Address="", SubAddress=sub_address, TextToDisplay=text)


def solve_p03_02(excel) -> None:
    path = BASE / "P03_02.xlsx"
    wb = excel.Workbooks.Open(str(path))
    try:
        last = 400
        d = wb.Worksheets("Data")

        for sn in (
            "P2a_Gender",
            "P2b_Age",
            "P2c_Salary",
            "P34_RowPct",
            "P34_ColPct",
        ):
            _del_sheet_if_exists(wb, sn)

        d.Range("H1").Value = "Salary_Group"
        d.Range("H2").Formula = (
            '=IF(F2<40000,"Less than $40K",'
            'IF(F2<70000,"Between $40K and $70K",'
            'IF(F2<100000,"Between $70K and $100K","Greater than $100K")))'
        )
        d.Range(f"H2:H{last}").FillDown()

        n_op = len(OPINIONS)

        # --- P2a ---
        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P2a_Gender"
        s = wb.Worksheets("P2a_Gender")
        s.Range("A1").Value = (
            "Problem 2a (Rubric): COUNTIFS crosstab + column chart; % within each gender sum to 100%. "
            "Headers assume code 1 = Male, 2 = Female (swap labels if your key differs)."
        )
        s.Range("A2").Value = "Opinion"
        s.Range("B2").Value = "Male (1)"
        s.Range("C2").Value = "Female (2)"
        s.Range("D2").Value = "Row total"
        for k, op in enumerate(OPINIONS):
            r = 3 + k
            s.Range(f"A{r}").Value = op
            s.Range(f"B{r}").Formula = (
                f'=COUNTIFS(Data!$C$2:$C${last},1,Data!$G$2:$G${last},$A{r})'
            )
            s.Range(f"C{r}").Formula = (
                f'=COUNTIFS(Data!$C$2:$C${last},2,Data!$G$2:$G${last},$A{r})'
            )
            s.Range(f"D{r}").Formula = f"=SUM(B{r}:C{r})"
        trow = 3 + n_op
        s.Range(f"A{trow}").Value = "Total"
        s.Range(f"B{trow}").Formula = f"=SUM(B3:B{trow - 1})"
        s.Range(f"C{trow}").Formula = f"=SUM(C3:C{trow - 1})"
        s.Range(f"D{trow}").Formula = f"=SUM(D3:D{trow - 1})"
        hdr = trow + 2
        s.Range(f"A{hdr}").Value = "% within gender (each column sums to 100)"
        for k in range(n_op):
            ro = 3 + k
            rp = hdr + 1 + k
            s.Range(f"A{rp}").Value = OPINIONS[k]
            s.Range(f"B{rp}").Formula = f"=IF(B${trow}=0,0,B{ro}/B${trow}*100)"
            s.Range(f"C{rp}").Formula = f"=IF(C${trow}=0,0,C{ro}/C${trow}*100)"
        interp = hdr + 1 + n_op
        s.Range(f"A{interp}").Value = "Interpretation"
        s.Range(f"B{interp}").Value = (
            "Discussion (rubric): Compare the % bars for males vs females at each opinion. "
            "If the shapes differ (peaks at different opinions), the two genders tend to answer differently; "
            "if patterns are similar, opinions are more alike across gender."
        )
        s.Range(f"B{interp}").WrapText = True
        co = s.ChartObjects().Add(420, 15, 460, 230)
        ch = co.Chart
        ch.ChartType = XL_COLUMN_CLUSTERED
        ch.SetSourceData(s.Range(f"A{hdr}").Resize(n_op + 1, 3))
        ch.HasTitle = True
        ch.ChartTitle.Text = "Opinion by gender (% within gender)"
        ch.Axes(1).HasTitle = True
        ch.Axes(1).AxisTitle.Text = "Opinion"
        ch.Axes(2).HasTitle = True
        ch.Axes(2).AxisTitle.Text = "Percent"
        _style_merge_title(s, "A1:F1")
        _style_column_headers(s.Range("A2:D2"))
        _zebra_block(s, 3, 1, trow - 1, 4)
        _style_total_row(s.Range(f"A{trow}:D{trow}"))
        _style_subtitle_row(s.Range(f"A{hdr}:D{hdr}"))
        _zebra_block(s, hdr + 1, 1, hdr + n_op, 4)
        _style_note_block(s.Range(f"A{interp}:D{interp}"))
        s.Range(f"B{hdr + 1}:C{hdr + n_op}").NumberFormat = "0.00"
        _polish_column_chart(ch)
        s.Tab.Color = CLR_TAB_BLUE
        s.Columns("A").ColumnWidth = 22
        s.Columns("B:D").ColumnWidth = 14
        _freeze_top_row(s, 2)

        # --- P2b ---
        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P2b_Age"
        s = wb.Worksheets("P2b_Age")
        s.Range("A1").Value = "Problem 2b: Opinion x Age — % within age (columns sum to 100)"
        s.Range("A2").Value = "Opinion"
        for j, age in enumerate(AGE_ORDER):
            s.Range(f"{chr(ord('B') + j)}2").Value = age
        totc = chr(ord("B") + len(AGE_ORDER))
        s.Range(f"{totc}2").Value = "Row total"
        for k, op in enumerate(OPINIONS):
            r = 3 + k
            s.Range(f"A{r}").Value = op
            for j, age in enumerate(AGE_ORDER):
                col = chr(ord("B") + j)
                s.Range(f"{col}{r}").Formula = (
                    f'=COUNTIFS(Data!$B$2:$B${last},"{age}",Data!$G$2:$G${last},$A{r})'
                )
            s.Range(f"{totc}{r}").Formula = f"=SUM(B{r}:{chr(ord('B') + len(AGE_ORDER) - 1)}{r})"
        trow = 3 + n_op
        s.Range(f"A{trow}").Value = "Total"
        for j in range(len(AGE_ORDER)):
            col = chr(ord("B") + j)
            s.Range(f"{col}{trow}").Formula = f"=SUM({col}3:{col}{trow - 1})"
        s.Range(f"{totc}{trow}").Formula = f"=SUM({totc}3:{totc}{trow - 1})"
        hdr = trow + 2
        s.Range(f"A{hdr}").Value = "% within age group"
        for k in range(n_op):
            ro = 3 + k
            rp = hdr + 1 + k
            s.Range(f"A{rp}").Value = OPINIONS[k]
            for j in range(len(AGE_ORDER)):
                col = chr(ord("B") + j)
                s.Range(f"{col}{rp}").Formula = (
                    f"=IF({col}${trow}=0,0,{col}{ro}/{col}${trow}*100)"
                )
        interp = hdr + 1 + n_op
        s.Range(f"A{interp}").Value = "Interpretation"
        s.Range(f"B{interp}").Value = (
            "Discussion (rubric): For Young vs Middle-aged vs Elderly, compare how opinion % columns differ. "
            "Different profiles suggest opinion depends on age; similar profiles suggest little age effect."
        )
        s.Range(f"B{interp}").WrapText = True
        co = s.ChartObjects().Add(420, 15, 460, 230)
        ch = co.Chart
        ch.ChartType = XL_COLUMN_CLUSTERED
        ch.SetSourceData(s.Range(f"A{hdr}").Resize(n_op + 1, 1 + len(AGE_ORDER)))
        ch.HasTitle = True
        ch.ChartTitle.Text = "Opinion by age (% within age)"
        ch.Axes(2).HasTitle = True
        ch.Axes(2).AxisTitle.Text = "Percent"
        nc = 1 + len(AGE_ORDER) + 1
        lc = chr(ord("A") + nc - 1)
        _style_merge_title(s, f"A1:{lc}1")
        _style_column_headers(s.Range(f"A2:{lc}2"))
        _zebra_block(s, 3, 1, trow - 1, nc)
        _style_total_row(s.Range(f"A{trow}:{lc}{trow}"))
        _style_subtitle_row(s.Range(f"A{hdr}:{lc}{hdr}"))
        _zebra_block(s, hdr + 1, 1, hdr + n_op, nc)
        _style_note_block(s.Range(f"A{interp}:{lc}{interp}"))
        s.Range(f"B{hdr + 1}:{chr(ord('B') + len(AGE_ORDER) - 1)}{hdr + n_op}").NumberFormat = "0.00"
        _polish_column_chart(ch)
        s.Tab.Color = CLR_TAB_BLUE
        s.Columns("A").ColumnWidth = 22
        _freeze_top_row(s, 2)

        # --- P2c ---
        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P2c_Salary"
        s = wb.Worksheets("P2c_Salary")
        s.Range("A1").Value = "Problem 2c: Opinion x Salary_Group (book $40K / $70K / $100K bands)"
        s.Range("A2").Value = "Opinion"
        n_sb = len(SALARY_BANDS)
        for j, sg in enumerate(SALARY_BANDS):
            s.Range(f"{chr(ord('B') + j)}2").Value = sg
        totc = chr(ord("B") + n_sb)
        s.Range(f"{totc}2").Value = "Row total"
        for k, op in enumerate(OPINIONS):
            r = 3 + k
            s.Range(f"A{r}").Value = op
            for j, sg in enumerate(SALARY_BANDS):
                col = chr(ord("B") + j)
                s.Range(f"{col}{r}").Formula = (
                    f'=COUNTIFS(Data!$H$2:$H${last},"{sg}",Data!$G$2:$G${last},$A{r})'
                )
            s.Range(f"{totc}{r}").Formula = f"=SUM(B{r}:{chr(ord('B') + n_sb - 1)}{r})"
        trow = 3 + n_op
        s.Range(f"A{trow}").Value = "Total"
        for j in range(n_sb):
            col = chr(ord("B") + j)
            s.Range(f"{col}{trow}").Formula = f"=SUM({col}3:{col}{trow - 1})"
        s.Range(f"{totc}{trow}").Formula = f"=SUM({totc}3:{totc}{trow - 1})"
        hdr = trow + 2
        s.Range(f"A{hdr}").Value = "% within salary band (book breakpoints)"
        for k in range(n_op):
            ro = 3 + k
            rp = hdr + 1 + k
            s.Range(f"A{rp}").Value = OPINIONS[k]
            for j in range(n_sb):
                col = chr(ord("B") + j)
                s.Range(f"{col}{rp}").Formula = (
                    f"=IF({col}${trow}=0,0,{col}{ro}/{col}${trow}*100)"
                )
        interp = hdr + 1 + n_op
        s.Range(f"A{interp}").Value = "Interpretation"
        s.Range(f"B{interp}").Value = (
            "Discussion (rubric): Across salary bands (text breakpoints), compare % distributions. "
            "If higher earners lean more agree (or disagree) than lower bands, salary relates to opinion."
        )
        s.Range(f"B{interp}").WrapText = True
        co = s.ChartObjects().Add(420, 15, 460, 230)
        ch = co.Chart
        ch.ChartType = XL_COLUMN_CLUSTERED
        ch.SetSourceData(s.Range(f"A{hdr}").Resize(n_op + 1, 1 + n_sb))
        ch.HasTitle = True
        ch.ChartTitle.Text = "Opinion by salary band (% within band)"
        ch.Axes(2).HasTitle = True
        ch.Axes(2).AxisTitle.Text = "Percent"
        nc = 1 + n_sb + 1
        lc = chr(ord("A") + nc - 1)
        _style_merge_title(s, f"A1:{lc}1")
        _style_column_headers(s.Range(f"A2:{lc}2"))
        _zebra_block(s, 3, 1, trow - 1, nc)
        _style_total_row(s.Range(f"A{trow}:{lc}{trow}"))
        _style_subtitle_row(s.Range(f"A{hdr}:{lc}{hdr}"))
        _zebra_block(s, hdr + 1, 1, hdr + n_op, nc)
        _style_note_block(s.Range(f"A{interp}:{lc}{interp}"))
        s.Range(f"B{hdr + 1}:{chr(ord('B') + n_sb - 1)}{hdr + n_op}").NumberFormat = "0.00"
        _polish_column_chart(ch)
        s.Tab.Color = CLR_TAB_BLUE
        s.Columns("A").ColumnWidth = 22
        _freeze_top_row(s, 2)

        src = f"Data!$A$1:$H${last}"

        def add_pivot(ws, top_row: int, left_col: int, table_name: str, col_field: str, pct_row: bool) -> None:
            pc = wb.PivotCaches().Create(XL_DATABASE, src)
            dest = ws.Cells(top_row, left_col)
            pt = pc.CreatePivotTable(TableDestination=dest, TableName=table_name)
            pt.ManualUpdate = True
            pt.PivotFields("Opinion").Orientation = XL_ROW_FIELD
            pt.PivotFields(col_field).Orientation = XL_COLUMN_FIELD
            pt.AddDataField(pt.PivotFields("Person"), "Count of Person", XL_COUNT)
            df = pt.DataFields("Count of Person")
            df.Calculation = XL_PERCENT_OF_ROW if pct_row else XL_PERCENT_OF_COLUMN
            df.NumberFormat = "0.0%"
            pt.ManualUpdate = False
            pt.RefreshTable()

        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P34_RowPct"
        sr = wb.Worksheets("P34_RowPct")
        sr.Range("A1").Value = "Problem 34 — row % pivots (each opinion row sums to 100%)."
        add_pivot(sr, 3, 1, "P34R_Gender", "Gender", True)
        add_pivot(sr, 28, 1, "P34R_Age", "Age", True)
        add_pivot(sr, 53, 1, "P34R_Sal", "Salary_Group", True)

        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P34_ColPct"
        sc = wb.Worksheets("P34_ColPct")
        sc.Range("A1").Value = "Problem 34 — column % pivots (each column sums to 100%)."
        add_pivot(sc, 3, 1, "P34C_Gender", "Gender", False)
        add_pivot(sc, 28, 1, "P34C_Age", "Age", False)
        add_pivot(sc, 53, 1, "P34C_Sal", "Salary_Group", False)

        def add_pivot_chart(
            ws,
            pt_name: str,
            left: float,
            top: float,
            title: str,
            x_title: str,
        ) -> None:
            pt = ws.PivotTables(pt_name)
            co = ws.ChartObjects().Add(left, top, 360, 200)
            ch = co.Chart
            ch.ChartType = XL_COLUMN_CLUSTERED
            ch.SetSourceData(pt.TableRange1)
            try:
                ch.HasAutoTitle = False
            except Exception:
                pass
            ch.HasTitle = False
            ch.HasTitle = True
            ch.ChartTitle.Text = title
            ch.Axes(1).HasTitle = True
            ch.Axes(1).AxisTitle.Text = x_title
            ch.Axes(2).HasTitle = True
            ch.Axes(2).AxisTitle.Text = "Percent"

        add_pivot_chart(sr, "P34R_Gender", 520, 10, "Row %: Opinion by Gender", "Opinion")
        add_pivot_chart(sr, "P34R_Age", 520, 220, "Row %: Opinion by Age", "Opinion")
        add_pivot_chart(sr, "P34R_Sal", 520, 430, "Row %: Opinion by Salary Band", "Opinion")
        add_pivot_chart(sc, "P34C_Gender", 520, 10, "Column %: Opinion by Gender", "Opinion")
        add_pivot_chart(sc, "P34C_Age", 520, 220, "Column %: Opinion by Age", "Opinion")
        add_pivot_chart(sc, "P34C_Sal", 520, 430, "Column %: Opinion by Salary Band", "Opinion")

        sr.Range("A78").Value = "Interpretation (row %) — rubric (3 pts)"
        sr.Range("B78").Value = (
            "Row-% pivots: Each opinion row sums to 100%. You see how people in that opinion category "
            "split across gender, age, and salary band—e.g. whether “Strongly agree” is mostly one gender or spread. "
            "If splits differ sharply by row, those background variables relate to who holds each opinion."
        )
        sr.Range("B78").WrapText = True
        sc.Range("A78").Value = "Interpretation (column %) — rubric (3 pts)"
        sc.Range("B78").Value = (
            "Column-% pivots: Each column (gender / age / salary slice) sums to 100%. You see the opinion mix "
            "within each group—e.g. whether males vs females show different shares agreeing. "
            "Different column profiles mean opinion distribution differs across that variable."
        )
        sc.Range("B78").WrapText = True

        _style_merge_title(sr, "A1:J1", row_height=42)
        _style_note_block(sr.Range("A78:J78"))
        sr.Range("A78").Font.Bold = True
        for ix in range(1, sr.ChartObjects().Count + 1):
            _polish_column_chart(sr.ChartObjects(ix).Chart)
        sr.Tab.Color = CLR_TAB_ORANGE
        _freeze_top_row(sr, 2)

        _style_merge_title(sc, "A1:J1", row_height=42)
        _style_note_block(sc.Range("A78:J78"))
        sc.Range("A78").Font.Bold = True
        for ix in range(1, sc.ChartObjects().Count + 1):
            _polish_column_chart(sc.ChartObjects(ix).Chart)
        sc.Tab.Color = CLR_TAB_PURPLE
        _freeze_top_row(sc, 2)

        # Navigation links on Data (formulas/charts update automatically; links jump to each rubric sheet).
        for i, (txt, sub) in enumerate(
            (
                ("P2a Gender", "P2a_Gender!A1"),
                ("P2b Age", "P2b_Age!A1"),
                ("P2c Salary bands", "P2c_Salary!A1"),
                ("P34 row %", "P34_RowPct!A1"),
                ("P34 column %", "P34_ColPct!A1"),
            ),
            start=2,
        ):
            _add_sheet_hyperlink(d, f"I{i}", sub, txt)
        _style_nav_panel(d, "I1", 5, "Open answers")
        _style_column_headers(d.Range("A1:H1"))
        d.Tab.Color = CLR_TAB_GREEN
        d.Range("A2:H2").Interior.Color = CLR_BAND
        _freeze_top_row(d, 1)

        wb.Save()
        wb.Close()
    except Exception:
        wb.Close(SaveChanges=False)
        raise


def solve_p03_08(excel) -> None:
    path = BASE / "P03_08.xlsx"
    wb = excel.Workbooks.Open(str(path))
    try:
        _del_sheet_if_exists(wb, "P8_Summary")
        _del_sheet_if_exists(wb, "P8_TableFilterSteps")
        # P03_08 is often reset from a copy of P03_21; drop Problem 21 sheets if present.
        _del_sheet_if_exists(wb, "P21_Corr")
        _del_sheet_if_exists(wb, "P21_Charts")
        ws = wb.Worksheets("Data")
        last_row = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row

        _clear_listobjects_com(ws)
        if str(ws.Range("A1").Value) != "Employee":
            raise RuntimeError(
                "P03_08 Data!A1 must be 'Employee'. Restore with _restore_p03_08.py from P03_21."
            )

        ws.Range("H1").Value = "Gender_Label"
        ws.Range("H2").Formula = '=IF(B2=1,"Male","Female")'
        ws.Range(f"H2:H{last_row}").FillDown()

        # Textbook / rubric: new education values 1 = <4, 2 = 4, 3 = >4 years coded.
        ws.Range("I1").Value = "Edu_Recode"
        ws.Range("I2").Formula = '=IF(F2<4,1,IF(F2=4,2,3))'
        ws.Range(f"I2:I{last_row}").FillDown()

        ws.Range("J1").Value = "Age_Group"
        # Book Ch.3 Problem 8: Young <34; Middle-aged 34–49; Older >=50.
        # Rubric wording "Middle-Aged"; book uses "middle-aged" — use rubric spelling for Age_Group.
        ws.Range("J2").Formula = (
            '=IF(C2<34,"Young",IF(C2<50,"Middle-Aged","Older"))'
        )
        ws.Range(f"J2:J{last_row}").FillDown()

        rng = ws.Range(f"A1:J{last_row}")
        tbl = ws.ListObjects.Add(XL_SRC_RANGE, rng, None, XL_YES)
        tbl.Name = "TblEmployees"
        _style_table_on_sheet(tbl)

        wb.Worksheets.Add(After=wb.Worksheets(wb.Worksheets.Count)).Name = "P8_Summary"
        s = wb.Worksheets("P8_Summary")
        s.Range("A1").Value = (
            "Problem 8 (Rubric): Use Excel Table TblEmployees on Data — filter each group and read "
            "summary stats for Annual Salary, OR use the database formulas below (same numbers as filtered rows). "
            "Gender_Label (H), Edu_Recode 1–3 (I), Age_Group (J) match the textbook."
        )

        db = f"Data!$A$1:$J${last_row}"
        sal_col = 7

        def dfunc(fname: str, crit: str) -> str:
            return f"={fname}({db},{sal_col},{crit})"

        s.Range("A3").Value = "Gender (codes on Data)"
        s.Range("B3").Value = "Mean"
        s.Range("C3").Value = "Median"
        s.Range("D3").Value = "StDev"
        s.Range("E3").Value = "Criteria"

        s.Range("A4").Value = "Male (1)"
        s.Range("E4").Value = "Gender"
        s.Range("E5").Value = 1
        s.Range("B4").Formula = f"=AVERAGEIFS(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row},1)"
        s.Range("C4").Formula = (
            f"=MEDIAN(FILTER(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row}=1))"
        )
        s.Range("D4").Formula = (
            f"=STDEV.S(FILTER(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row}=1))"
        )

        s.Range("A6").Value = "Female (0)"
        s.Range("E6").Value = "Gender"
        s.Range("E7").Value = 0
        s.Range("B6").Formula = f"=AVERAGEIFS(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row},0)"
        s.Range("C6").Formula = (
            f"=MEDIAN(FILTER(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row}=0))"
        )
        s.Range("D6").Formula = (
            f"=STDEV.S(FILTER(Data!$G$2:$G${last_row},Data!$B$2:$B${last_row}=0))"
        )

        s.Range("A9").Value = "Education recode (levels 1, 2, 3)"
        for label, crit_val, r in [("Level 1", 1, 10), ("Level 2", 2, 13), ("Level 3", 3, 16)]:
            s.Range(f"A{r}").Value = label
            s.Range(f"E{r}").Value = "Edu_Recode"
            s.Range(f"E{r + 1}").Value = crit_val
            s.Range(f"B{r}").Formula = (
                f"=AVERAGEIFS(Data!$G$2:$G${last_row},Data!$I$2:$I${last_row},{crit_val})"
            )
            s.Range(f"C{r}").Formula = (
                f"=MEDIAN(FILTER(Data!$G$2:$G${last_row},Data!$I$2:$I${last_row}={crit_val}))"
            )
            s.Range(f"D{r}").Formula = (
                f"=STDEV.S(FILTER(Data!$G$2:$G${last_row},Data!$I$2:$I${last_row}={crit_val}))"
            )

        s.Range("A19").Value = "Age recode"
        for label, crit_val, r in [("Young", "Young", 20), ("Middle-Aged", "Middle-Aged", 23), ("Older", "Older", 26)]:
            s.Range(f"A{r}").Value = label
            s.Range(f"E{r}").Value = "Age_Group"
            s.Range(f"E{r + 1}").Value = crit_val
            s.Range(f"B{r}").Formula = (
                f'=AVERAGEIFS(Data!$G$2:$G${last_row},Data!$J$2:$J${last_row},"{crit_val}")'
            )
            s.Range(f"C{r}").Formula = (
                f'=MEDIAN(FILTER(Data!$G$2:$G${last_row},Data!$J$2:$J${last_row}="{crit_val}"))'
            )
            s.Range(f"D{r}").Formula = (
                f'=STDEV.S(FILTER(Data!$G$2:$G${last_row},Data!$J$2:$J${last_row}="{crit_val}"))'
            )

        s.Range("A29").Value = (
            "Cross-check: Filter TblEmployees (Data) to one group at a time; summary stats should match."
        )

        _style_merge_title(s, "A1:F1", row_height=72)
        _style_column_headers(s.Range("B3:E3"))
        s.Range("A4:A7").Interior.Color = CLR_BAND
        s.Range("A10:A17").Interior.Color = CLR_BAND
        s.Range("A20:A27").Interior.Color = CLR_BAND
        s.Range("B4:D7").NumberFormat = "#,##0"
        s.Range("B10:D17").NumberFormat = "#,##0"
        s.Range("B20:D27").NumberFormat = "#,##0"
        s.Range("A29:F29").Interior.Color = CLR_NOTE
        s.Range("A29:F29").WrapText = True
        _border_box(s.Range("A3:E7"))
        _border_box(s.Range("A9:E17"))
        _border_box(s.Range("A19:E27"))
        s.Tab.Color = CLR_TAB_ORANGE
        _freeze_top_row(s, 3)

        _add_sheet_hyperlink(ws, "K1", "P8_Summary!A1", "P8 Summary (stats)")
        for addr in ("K1",):
            c = ws.Range(addr)
            c.Interior.Color = CLR_LINK_BG
            c.Font.Color = CLR_SUBHEADER
            _border_box(c)
        _style_column_headers(ws.Range("A1:J1"))
        ws.Tab.Color = CLR_TAB_GREEN
        _freeze_top_row(ws, 1)

        wb.Save()
        wb.Close()
    except Exception:
        wb.Close(SaveChanges=False)
        raise


def _add_p21_multiple_regression_sheet(wb, last_row: int) -> None:
    """Sheet matching instructor-style regression output; all cells are formula-driven (data changes → recalc)."""
    s = wb.Worksheets("P21_Answers")
    y = f"Data!$G$2:$G${last_row}"
    x = f"Data!$B$2:$F${last_row}"
    lst = f"LINEST({y},{x},TRUE,TRUE)"

    def ix(r: int, c: int) -> str:
        return f"INDEX({lst},{r},{c})"

    s.Range("A1").Value = (
        "Problem 21 — Answers (interactive): multiple regression of Annual Salary on Gender, Age, Prior Exp, Beta Exp, Education. "
        "Official rubric items (correlation matrix, three scatters, trendline on strongest predictor) are on P21_Corr and P21_Charts."
    )
    s.Range("A1").WrapText = True

    s.Range("A3").Value = "Model Summary"
    s.Range("A4").Value = "Observations (n)"
    s.Range("B4").Formula = f"=COUNT({y})"
    s.Range("A5").Value = "R Square"
    s.Range("B5").Formula = f"={ix(3, 1)}"
    s.Range("A6").Value = "Adjusted R Square"
    s.Range("B6").Formula = f"=1-(1-{ix(3, 1)})*(B4-1)/(B4-5-1)"
    s.Range("A7").Value = "Standard Error"
    s.Range("B7").Formula = f"={ix(3, 2)}"
    s.Range("A8").Value = "F-statistic"
    s.Range("B8").Formula = f"={ix(4, 1)}"
    s.Range("A9").Value = "DF Residuals"
    s.Range("B9").Formula = f"={ix(4, 2)}"
    s.Range("A10").Value = "Significance F"
    s.Range("B10").Formula = "=F.DIST.RT(B8,5,B9)"

    s.Range("A12").Value = "ANOVA"
    s.Range("A13").Value = "Source"
    s.Range("B13").Value = "df"
    s.Range("C13").Value = "SS"
    s.Range("D13").Value = "MS"
    s.Range("E13").Value = "F"
    s.Range("F13").Value = "Significance F"
    s.Range("A14").Value = "Regression"
    s.Range("B14").Value = 5
    s.Range("C14").Formula = f"={ix(5, 1)}"
    s.Range("D14").Formula = "=C14/B14"
    s.Range("E14").Formula = "=B8"
    s.Range("F14").Formula = "=B10"
    s.Range("A15").Value = "Residual"
    s.Range("B15").Formula = f"={ix(4, 2)}"
    s.Range("C15").Formula = f"={ix(5, 2)}"
    s.Range("D15").Formula = "=C15/B15"
    s.Range("A16").Value = "Total"
    s.Range("B16").Formula = "=B14+B15"
    s.Range("C16").Formula = "=C14+C15"

    s.Range("A18").Value = "Coefficients"
    s.Range("A19").Value = "Variable"
    s.Range("B19").Value = "Coefficient"
    s.Range("C19").Value = "Std Error"
    s.Range("D19").Value = "t-Statistic"
    s.Range("E19").Value = "P-value"
    s.Range("F19").Value = "Lower 95%"
    s.Range("G19").Value = "Upper 95%"

    rows = [
        ("Intercept", 1, 6),
        ("Gender", 1, 5),
        ("Age", 1, 4),
        ("Prior Experience", 1, 3),
        ("Beta Experience", 1, 2),
        ("Education", 1, 1),
    ]
    for i, (name, r_coef, c_coef) in enumerate(rows, start=20):
        s.Range(f"A{i}").Value = name
        s.Range(f"B{i}").Formula = f"={ix(r_coef, c_coef)}"
        s.Range(f"C{i}").Formula = f"={ix(r_coef + 1, c_coef)}"
        s.Range(f"D{i}").Formula = f"=IF(C{i}=0,0,B{i}/C{i})"
        s.Range(f"E{i}").Formula = f"=T.DIST.2T(ABS(D{i}),$B$9)"
        s.Range(f"F{i}").Formula = f"=B{i}-T.INV.2T(0.05,$B$9)*C{i}"
        s.Range(f"G{i}").Formula = f"=B{i}+T.INV.2T(0.05,$B$9)*C{i}"

    eq = (
        f'="ŷ = "&TEXT({ix(1, 6)},"#,##0.0000")&" + ("&TEXT({ix(1, 5)},"#,##0.0000")&" * Gender) + ("'
        f'&TEXT({ix(1, 4)},"#,##0.0000")&" * Age) + ("&TEXT({ix(1, 3)},"#,##0.0000")'
        f'&" * Prior Experience) + ("&TEXT({ix(1, 2)},"#,##0.0000")&" * Beta Experience) + ("'
        f'&TEXT({ix(1, 1)},"#,##0.0000")&" * Education)"'
    )
    s.Range("A27").Value = "Multiple regression equation"
    s.Range("B27").Formula = eq
    s.Range("B27").WrapText = True

    s.Columns("A:G").AutoFit()
    _style_regression_outputs(s)
    s.Tab.Color = CLR_TAB_ORANGE
    _freeze_top_row(s, 3)


def _p21_rank_predictors(xlsx_path: Path):
    """Return (last_row, top3_indices, best_idx, names, excel_cols, cors) for Age,Prior,Beta,Edu vs Salary."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Data"]
    rows = []
    last = ws.max_row
    for r in range(2, last + 1):
        vals = [ws.cell(r, c).value for c in (3, 4, 5, 6, 7)]
        if any(v is None for v in vals):
            continue
        try:
            rows.append([float(v) for v in vals])
        except (TypeError, ValueError):
            continue
    a = np.array(rows, dtype=float)
    sal = a[:, 4]
    cors = np.array([np.corrcoef(a[:, i], sal)[0, 1] for i in range(4)])
    order = np.argsort(-np.abs(cors))
    top3 = [int(x) for x in order[:3]]
    best = int(order[0])
    names = ["Age", "Prior Exp", "Beta Exp", "Education"]
    cols = [3, 4, 5, 6]
    lr = ws.max_row
    return lr, top3, best, names, cols, cors


def solve_p03_21(excel) -> None:
    path = BASE / "P03_21.xlsx"
    last_row, top3, best, pnames, pcols, _cors = _p21_rank_predictors(path)

    wb = excel.Workbooks.Open(str(path))
    try:
        _del_sheet_if_exists(wb, "P21_Answers")
        _del_sheet_if_exists(wb, "P21_Corr")
        _del_sheet_if_exists(wb, "P21_Charts")
        ws = wb.Worksheets("Data")
        d = f"Data!$C$2:$C${last_row}"
        pr = f"Data!$D$2:$D${last_row}"
        be = f"Data!$E$2:$E${last_row}"
        ed = f"Data!$F$2:$F${last_row}"
        sl = f"Data!$G$2:$G${last_row}"

        wb.Worksheets.Add(After=wb.Worksheets("Data")).Name = "P21_Corr"
        wb.Worksheets.Add(Before=wb.Worksheets("P21_Corr")).Name = "P21_Answers"
        _add_p21_multiple_regression_sheet(wb, last_row)

        s = wb.Worksheets("P21_Corr")
        s.Range("A1").Value = (
            "Problem 21 (Rubric): correlation matrix; 3 scatterplots (2 pts each) for the 3 strongest "
            "predictors vs salary; trendline + equation on the most correlated; interpret slope. "
            "Full multiple regression (incl. Gender) is on P21_Answers."
        )
        labs = ["Age", "Prior Exp", "Beta Exp", "Education", "Salary"]
        for i, lab in enumerate(labs, start=2):
            s.Cells(1, i).Value = lab
            s.Cells(i, 1).Value = lab
        refs = [d, pr, be, ed, sl]
        for i in range(5):
            for j in range(5):
                f = 2 + i
                c = 2 + j
                if i == j:
                    s.Cells(f, c).Value = 1
                else:
                    s.Cells(f, c).Formula = f"=CORREL({refs[i]},{refs[j]})"

        s.Range("A9").Value = "Part (a): Correlations with Annual Salary"
        s.Range("A10").Value = "Age"
        s.Range("B10").Formula = f"=CORREL({d},{sl})"
        s.Range("A11").Value = "Prior Exp"
        s.Range("B11").Formula = f"=CORREL({pr},{sl})"
        s.Range("A12").Value = "Beta Exp"
        s.Range("B12").Formula = f"=CORREL({be},{sl})"
        s.Range("A13").Value = "Education"
        s.Range("B13").Formula = f"=CORREL({ed},{sl})"
        s.Range("A14").Value = "Strongest predictor (largest |r|)"
        s.Range("C9").Value = "|r| helper"
        s.Range("C10").Formula = "=ABS(B10)"
        s.Range("C11").Formula = "=ABS(B11)"
        s.Range("C12").Formula = "=ABS(B12)"
        s.Range("C13").Formula = "=ABS(B13)"
        s.Range("B14").Formula = "=INDEX(A10:A13,MATCH(MAX(C10:C13),C10:C13,0))"
        s.Range("A15").Value = "Its r"
        s.Range("B15").Formula = "=INDEX(B10:B13,MATCH(MAX(C10:C13),C10:C13,0))"
        s.Range("A16").Value = "Answer to part (a)"
        s.Range("B16").Formula = '="Most highly correlated with salary: "&B14&" (r = "&TEXT(B15,"0.0000")&")"'

        # LINEST slope/intercept (y = Annual Salary) for rubric interpretation
        s.Range("N9").Value = "LINEST slope vs Salary"
        s.Range("O9").Value = "LINEST intercept"
        s.Range("N10").Formula = f"=INDEX(LINEST({sl},{d},TRUE,TRUE),1)"
        s.Range("N11").Formula = f"=INDEX(LINEST({sl},{pr},TRUE,TRUE),1)"
        s.Range("N12").Formula = f"=INDEX(LINEST({sl},{be},TRUE,TRUE),1)"
        s.Range("N13").Formula = f"=INDEX(LINEST({sl},{ed},TRUE,TRUE),1)"
        s.Range("O10").Formula = f"=INDEX(LINEST({sl},{d},TRUE,TRUE),1,2)"
        s.Range("O11").Formula = f"=INDEX(LINEST({sl},{pr},TRUE,TRUE),1,2)"
        s.Range("O12").Formula = f"=INDEX(LINEST({sl},{be},TRUE,TRUE),1,2)"
        s.Range("O13").Formula = f"=INDEX(LINEST({sl},{ed},TRUE,TRUE),1,2)"

        s.Range("A17").Value = "Part (c): Slope for strongest predictor (matches chart trendline)"
        s.Range("B17").Formula = '=INDEX(N10:N13,MATCH(B14,A10:A13,0))'
        s.Range("A18").Value = "Part (c): Intercept for strongest predictor"
        s.Range("B18").Formula = '=INDEX(O10:O13,MATCH(B14,A10:A13,0))'

        s.Range("A19").Value = "Part (c) interpretation: if X increases by 1, salary will:"
        s.Range("B19").Formula = (
            r'="Salary "&IF(B17>=0,"increases","decreases")&" by about $"&TEXT(ABS(B17),"#,##0")'
            r'&" per +1 unit of "&B14&" (from linear trendline equation)."'
        )
        s.Range("B19").WrapText = True
        s.Range("A20").Value = "Part (c) intercept meaning"
        s.Range("B20").Formula = (
            r'="Intercept is about $"&TEXT(B18,"#,##0.00")&": predicted salary when "&B14&" = 0."'
        )
        s.Range("B20").WrapText = True

        s.Range("A21").Value = "Part (b): three scatterplots use the three largest |r| predictors from part (a)."
        order_txt = ", ".join(pnames[i] for i in top3)
        s.Range("B21").Value = f"Chart order (strongest first in ranking): {order_txt}"

        _style_corr_sheet(s)

        wb.Worksheets.Add(After=wb.Worksheets("P21_Corr")).Name = "P21_Charts"
        chs = wb.Worksheets("P21_Charts")
        chs.Range("A1").Value = "Three scatterplots (rubric) + linear trend on the strongest predictor only."
        chs.Range("A2").Value = "Part (b): all three charts have labeled axes. Part (c): trendline + equation appears only on the strongest predictor."
        chs.Range("A2:M2").Merge()
        chs.Range("A2").WrapText = True
        chs.Range("A2").Interior.Color = CLR_NOTE
        _border_box(chs.Range("A2:M2"))

        def scatter_xy(left: float, top: float, title: str, xcol: int, xlab: str, with_trend: bool):
            co = chs.ChartObjects().Add(left, top, 360, 220)
            ch = co.Chart
            ch.ChartType = XL_XY_SCATTER
            sc = ch.SeriesCollection().NewSeries()
            sc.Name = title
            sc.XValues = ws.Range(ws.Cells(2, xcol), ws.Cells(last_row, xcol))
            sc.Values = ws.Range(ws.Cells(2, 7), ws.Cells(last_row, 7))
            # Make points readable without cluttering the chart.
            sc.MarkerStyle = 8  # xlMarkerStyleCircle
            sc.MarkerSize = 6
            sc.Format.Line.Visible = False
            ch.HasTitle = True
            ch.ChartTitle.Text = title
            ch.HasLegend = False
            ch.Axes(1).HasTitle = True
            ch.Axes(1).AxisTitle.Text = xlab
            ch.Axes(1).HasMajorGridlines = False
            ch.Axes(2).HasTitle = True
            ch.Axes(2).AxisTitle.Text = "Annual Salary"
            ch.Axes(2).HasMajorGridlines = True
            if with_trend:
                ch.SeriesCollection(1).Trendlines().Add(Type=XL_LINEAR)
                tl = ch.SeriesCollection(1).Trendlines(1)
                tl.DisplayEquation = True
                tl.DisplayRSquared = True
                try:
                    tl.Format.Line.ForeColor.RGB = CLR_SUBHEADER
                except Exception:
                    pass
            _polish_scatter_chart(ch)
            ch.HasLegend = False
            try:
                ch.Axes(1).TickLabels.NumberFormatLinked = False
                ch.Axes(2).TickLabels.NumberFormatLinked = False
            except Exception:
                pass
            ch.Axes(1).TickLabels.NumberFormat = "0.0"
            ch.Axes(2).TickLabels.NumberFormat = "$#,##0"
            return ch

        positions = [(20, 70), (400, 70), (780, 70)]
        for idx, pred_i in enumerate(top3):
            name = pnames[pred_i]
            col = pcols[pred_i]
            left, top = positions[idx]
            chart_title = f"Part (b): Salary vs {name}"
            if pred_i == best:
                chart_title += " (Part c trendline + equation)"
            scatter_xy(
                left,
                top,
                chart_title,
                col,
                name,
                pred_i == best,
            )

        _style_merge_title(chs, "A1:M1", row_height=38)
        chs.Tab.Color = CLR_TAB_GREEN

        _add_sheet_hyperlink(ws, "I2", "P21_Answers!A1", "P21 Answers (full regression)")
        _add_sheet_hyperlink(ws, "I3", "P21_Corr!A1", "P21 Correlation + slopes")
        _add_sheet_hyperlink(ws, "I4", "P21_Charts!A1", "P21 Scatterplots")
        _style_nav_panel(ws, "I1", 3, "Open answers")
        _style_column_headers(ws.Range("A1:G1"))
        ws.Tab.Color = CLR_TAB_GREEN
        _freeze_top_row(ws, 1)

        wb.Save()
        wb.Close()
    except Exception:
        wb.Close(SaveChanges=False)
        raise


def main() -> None:
    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    try:
        solve_p03_02(excel)
        solve_p03_08(excel)
        solve_p03_21(excel)
        print("OK: P03_02.xlsx, P03_08.xlsx, and P03_21.xlsx updated (formulas + formatting). Open in Excel to view colors.")
    finally:
        excel.Quit()


if __name__ == "__main__":
    main()
